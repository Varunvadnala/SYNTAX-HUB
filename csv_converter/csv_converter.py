"""
CSV to Excel Converter
Usage: python csv_to_excel.py --input data.csv --output result.xlsx
"""

import argparse
import logging
import os
import sys

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# --- Logging Setup ---
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  [%(levelname)s]  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


# --- Step 1: Read CSV ---
def read_csv(filepath):
    if not os.path.exists(filepath):
        log.error(f"File not found: '{filepath}'")
        sys.exit(1)

    if not filepath.lower().endswith(".csv"):
        log.error(f"Expected a .csv file, got: '{filepath}'")
        sys.exit(1)

    try:
        df = pd.read_csv(filepath)
        log.info(f"Loaded '{filepath}' → {len(df)} rows, {len(df.columns)} columns")
        return df
    except pd.errors.EmptyDataError:
        log.error("The CSV file is empty.")
        sys.exit(1)
    except Exception as e:
        log.error(f"Could not read file: {e}")
        sys.exit(1)


# --- Step 2: Clean Data ---
def clean_data(df, fill_value, column_renames):
    # Clean column names
    df.columns = [col.strip() for col in df.columns]

    # Rename columns
    if column_renames:
        df = df.rename(columns=column_renames)
        for old, new in column_renames.items():
            log.info(f"Renamed column: '{old}' → '{new}'")

    # Auto-detect date columns
    for col in df.columns:
        if df[col].dtype == object:
            try:
                df[col] = pd.to_datetime(df[col], infer_datetime_format=True)
                log.info(f"Parsed '{col}' as dates.")
            except (ValueError, TypeError):
                pass

    # Fill missing values
    missing = df.isnull().sum().sum()
    if missing > 0:
        df = df.fillna(fill_value)
        log.info(f"Filled {missing} missing value(s) with '{fill_value}'.")

    # Strip spaces from text cells
    for col in df.select_dtypes(include=["object", "str"]).columns:
        df[col] = df[col].str.strip()

    log.info("Cleaning done.")
    return df


# --- Step 3: Export to Excel ---
def export_to_excel(df, output_path):
    if not output_path.lower().endswith(".xlsx"):
        output_path += ".xlsx"

    df.to_excel(output_path, index=False, sheet_name="Data")

    wb = load_workbook(output_path)
    ws = wb.active

    # Style header row
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", start_color="2E75B6")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Style data rows (alternating colours)
    light_blue = PatternFill("solid", start_color="DCE6F1")
    white = PatternFill("solid", start_color="FFFFFF")

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = light_blue if row_idx % 2 == 0 else white
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.font = Font(name="Arial", size=10)

    # Auto-fit column widths
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(cell.value)) for cell in col_cells if cell.value), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_path)
    log.info(f"✅ Saved to '{output_path}'")


# --- CLI ---
def parse_args():
    parser = argparse.ArgumentParser(description="Convert CSV to Excel")
    parser.add_argument("--input",  "-i", required=True, help="Input CSV file")
    parser.add_argument("--output", "-o", required=True, help="Output Excel file")
    parser.add_argument("--fill",   default="N/A",       help="Fill value for empty cells (default: N/A)")
    parser.add_argument("--rename", nargs="*", default=[],
                        metavar="OLD=NEW", help="Rename columns e.g. --rename 'Old=New'")
    return parser.parse_args()


def build_rename_dict(rename_list):
    renames = {}
    for item in rename_list:
        if "=" not in item:
            log.warning(f"Skipping bad --rename value: '{item}'")
            continue
        old, new = item.split("=", maxsplit=1)
        renames[old.strip()] = new.strip()
    return renames


# --- Main ---
def main():
    args = parse_args()
    log.info(f"Input: {args.input} | Output: {args.output} | Fill: '{args.fill}'")
    renames = build_rename_dict(args.rename)
    df = read_csv(args.input)
    df = clean_data(df, args.fill, renames)
    export_to_excel(df, args.output)

if __name__ == "__main__":
    main()