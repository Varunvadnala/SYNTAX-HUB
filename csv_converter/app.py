"""
CSV to Excel Converter - Web Application
A Flask-based web application for uploading, editing, and converting CSV to Excel
"""

from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import logging
import os
from datetime import datetime
from werkzeug.utils import secure_filename
import io
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import config

# Setup Flask app
app = Flask(__name__)
CORS(app)

# Load configuration
app.config['MAX_CONTENT_LENGTH'] = config.MAX_CONTENT_LENGTH
app.config['UPLOAD_FOLDER'] = config.UPLOAD_FOLDER

# Setup logging
logging.basicConfig(
    level=config.LOG_LEVEL,
    format=config.LOG_FORMAT,
    datefmt=config.LOG_DATE_FORMAT
)
log = logging.getLogger(__name__)

# Allowed extensions
ALLOWED_EXTENSIONS = config.ALLOWED_EXTENSIONS

# Store uploaded data in memory (for this session)
session_data = {}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def clean_data(df):
    """Clean and normalize data"""
    # Clean column names
    df.columns = [col.strip() for col in df.columns]
    
    # Auto-detect dates
    for col in df.columns:
        if df[col].dtype == object:
            try:
                df[col] = pd.to_datetime(df[col], infer_datetime_format=True)
            except (ValueError, TypeError):
                pass
    
    # Fill missing values
    df = df.fillna("N/A")
    
    # Strip spaces from text
    for col in df.select_dtypes(include=['object']).columns:
        df[col] = df[col].astype(str).str.strip()
    
    return df


def export_to_excel(df):
    """Export DataFrame to Excel in memory"""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Data')
        
        # Style the Excel file
        workbook = writer.book
        worksheet = writer.sheets['Data']
        
        # Style header row
        header_fill = PatternFill(start_color=config.HEADER_COLOR, end_color=config.HEADER_COLOR, fill_type="solid")
        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Style data rows
        light_blue = PatternFill(start_color=config.LIGHT_BLUE, end_color=config.LIGHT_BLUE, fill_type="solid")
        white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            fill = light_blue if row_idx % 2 == 0 else white
            for cell in row:
                cell.fill = fill
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = Font(name="Arial", size=10)
        
        # Auto-fit columns
        for col_idx, col_cells in enumerate(worksheet.columns, start=1):
            max_len = max((len(str(cell.value)) for cell in col_cells if cell.value), default=10)
            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)
        
        # Freeze header
        worksheet.freeze_panes = "A2"
    
    output.seek(0)
    return output


@app.route('/')
def index():
    """Render the main page"""
    return render_template('index.html')


@app.route('/api/upload', methods=['POST'])
def upload_file():
    """Handle CSV file upload"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'error': 'Only CSV files allowed'}), 400
        
        # Read CSV
        filename = secure_filename(file.filename)
        df = pd.read_csv(file)
        
        if df.empty:
            return jsonify({'success': False, 'error': 'CSV file is empty'}), 400
        
        # Clean data
        df = clean_data(df)
        
        # Store in session
        session_id = datetime.now().strftime("%Y%m%d%H%M%S")
        session_data[session_id] = df
        
        # Prepare response
        data = {
            'success': True,
            'message': f'Loaded {filename} - {len(df)} rows, {len(df.columns)} columns',
            'sessionId': session_id,
            'columns': df.columns.tolist(),
            'rows': df.values.tolist(),
            'stats': {
                'rows': len(df),
                'columns': len(df.columns),
                'filename': filename
            }
        }
        
        log.info(f"Uploaded {filename}: {len(df)} rows, {len(df.columns)} columns")
        return jsonify(data), 200
    
    except Exception as e:
        log.error(f"Upload error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/update-data', methods=['POST'])
def update_data():
    """Handle data updates from the web interface"""
    try:
        data = request.json
        session_id = data.get('sessionId')
        columns = data.get('columns')
        rows = data.get('rows')
        
        if session_id not in session_data:
            return jsonify({'success': False, 'error': 'Session not found'}), 400
        
        # Create new DataFrame from updated data
        df = pd.DataFrame(rows, columns=columns)
        session_data[session_id] = df
        
        log.info(f"Data updated: {len(df)} rows")
        return jsonify({'success': True, 'message': 'Data updated successfully'}), 200
    
    except Exception as e:
        log.error(f"Update error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/download-excel', methods=['POST'])
def download_excel():
    """Export data to Excel and download"""
    try:
        data = request.json
        session_id = data.get('sessionId')
        filename = data.get('filename', 'export.xlsx')
        
        if session_id not in session_data:
            return jsonify({'success': False, 'error': 'Session not found'}), 400
        
        df = session_data[session_id]
        excel_file = export_to_excel(df)
        
        log.info(f"Exporting to Excel: {filename}")
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            attachment_filename=filename
        )
    
    except Exception as e:
        log.error(f"Export error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/download-csv', methods=['POST'])
def download_csv():
    """Export data to CSV and download"""
    try:
        data = request.json
        session_id = data.get('sessionId')
        filename = data.get('filename', 'export.csv')
        
        if session_id not in session_data:
            return jsonify({'success': False, 'error': 'Session not found'}), 400
        
        df = session_data[session_id]
        csv_buffer = io.StringIO()
        df.to_csv(csv_buffer, index=False)
        csv_buffer.seek(0)
        
        csv_bytes = io.BytesIO(csv_buffer.getvalue().encode())
        log.info(f"Exporting to CSV: {filename}")
        
        return send_file(
            csv_bytes,
            mimetype='text/csv',
            as_attachment=True,
            attachment_filename=filename
        )
    
    except Exception as e:
        log.error(f"CSV export error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/add-row', methods=['POST'])
def add_row():
    """Add a new empty row"""
    try:
        data = request.json
        session_id = data.get('sessionId')
        
        if session_id not in session_data:
            return jsonify({'success': False, 'error': 'Session not found'}), 400
        
        df = session_data[session_id]
        new_row = {col: "" for col in df.columns}
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        session_data[session_id] = df
        
        return jsonify({'success': True, 'message': 'Row added', 'rowCount': len(df)}), 200
    
    except Exception as e:
        log.error(f"Add row error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/delete-row', methods=['POST'])
def delete_row():
    """Delete a row by index"""
    try:
        data = request.json
        session_id = data.get('sessionId')
        row_index = data.get('rowIndex')
        
        if session_id not in session_data:
            return jsonify({'success': False, 'error': 'Session not found'}), 400
        
        df = session_data[session_id]
        if 0 <= row_index < len(df):
            df = df.drop(row_index).reset_index(drop=True)
            session_data[session_id] = df
            return jsonify({'success': True, 'message': 'Row deleted', 'rowCount': len(df)}), 200
        
        return jsonify({'success': False, 'error': 'Invalid row index'}), 400
    
    except Exception as e:
        log.error(f"Delete row error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/add-column', methods=['POST'])
def add_column():
    """Add a new column"""
    try:
        data = request.json
        session_id = data.get('sessionId')
        column_name = data.get('columnName', 'New Column')
        
        if session_id not in session_data:
            return jsonify({'success': False, 'error': 'Session not found'}), 400
        
        df = session_data[session_id]
        if column_name not in df.columns:
            df[column_name] = ""
            session_data[session_id] = df
            return jsonify({'success': True, 'message': 'Column added', 'columns': df.columns.tolist()}), 200
        
        return jsonify({'success': False, 'error': 'Column already exists'}), 400
    
    except Exception as e:
        log.error(f"Add column error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


if __name__ == '__main__':
    log.info("Starting CSV to Excel Converter Web App...")
    app.run(debug=config.DEBUG, host=config.HOST, port=config.PORT)
